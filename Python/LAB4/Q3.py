from openpyxl import *

wb = load_workbook(filename = 'student.xlsx')
sheet = wb['Sheet1']

i = 2
while sheet.cell(row = i, column = 1).value is not None :
    midterm = sheet.cell(row = i, column = 3).value
    final = sheet.cell(row = i, column = 4).value
    hw = sheet.cell(row = i, column = 5).value
    attendance = sheet.cell(row = i, column = 6).value
    total = midterm * 0.3 + final * 0.35 + hw * 0.34 + attendance
    sheet.cell(row = i, column = 7, value = "{}".format(total))

    # total = 0
    # total += sheet.cell(row = i, column = 3).value * 0.3 # 총점 계산 
    # total += sheet.cell(row = i, column = 4).value * 0.35
    # total += sheet.cell(row = i, column = 5).value * 0.34
    # total += sheet.cell(row = i, column = 6).value
    # sheet.cell(row = i, column = 7, value = "{}".format(total))

    i += 1

wb.save("student.xlsx")
