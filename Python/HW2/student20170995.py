#!/usr/bin/python3
from openpyxl import *
import operator

wb = load_workbook(filename = 'student.xlsx')
sheet = wb['Sheet1']

i = 2
score = {}
while sheet.cell(row = i, column = 1).value is not None: 
    # { 행번호, 점수 } dictionary
    score[i] = float(sheet.cell(row = i, column = 7).value) 
    i += 1

scores = sorted(score.items(), key=operator.itemgetter(1), reverse = True)
# print(scores)
students = i - 2
rank = 1
for i in range(len(scores)) :
    idx = scores[i][0] # 실제 행 인덱스
    total = scores[i][1] # 현재 행 total(점수) 값

    j = i + 1
    if j < len(scores) :
        while total == scores[j][1] :
            j += 1

        if (j - i) >= 2:
            rank = j

    if total < 40 :
        sheet.cell(row = idx, column = 8, value = "F")
    elif rank <= students * 0.7 :
        if rank <= students * 0.3 :
            if rank <= students * 0.15 :
                sheet.cell(row = idx, column = 8, value = "A+")
            else :
                sheet.cell(row = idx, column = 8, value = "A0")
        else :
            if rank <= students * 0.5 :
                sheet.cell(row = idx, column = 8, value = "B+")
            else :
                sheet.cell(row = idx, column = 8, value = "B0")
    else :
        if rank <= students * 0.85 :
            sheet.cell(row = idx, column = 8, value = "C+")
        else :
            sheet.cell(row = idx, column = 8, value = "C0")
    rank += 1
    i += 1

wb.save('student.xlsx')
